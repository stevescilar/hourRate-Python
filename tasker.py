# work-life balancer
import time
import os
import platform
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook

# === Configuration ===
WORKING_HOURS = 11 * 3600   # 11hrs target
BREAK_TIME = 1 * 3600                # 1 hour
WARNING_TIME = 10 * 60               # 10 minutes
LOG_FILE_PATH = "work_timer_log.txt"
EXCEL_FILE_PATH = "work_timer_log.xlsx"

start_time = time.time()
start_time_str = datetime.now().strftime("%H:%M:%S")
start_date_str = datetime.now().strftime("%A, %B %d, %Y")


# === Utility Functions ===
def format_time(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"

def seconds_to_decimal_hours(seconds):
    return round(seconds / 3600, 1)


# === Logging Functions ===
def log_event(event):
    with open(LOG_FILE_PATH, "a") as f:
        f.write(f"{event}\n")

def open_log_file():
    system_name = platform.system()
    if system_name == "Windows":
        os.system(f'notepad.exe "{LOG_FILE_PATH}"')
    elif system_name == "Linux":
        os.system(f'xdg-open "{LOG_FILE_PATH}"')
    elif system_name == "Darwin":  # macOS
        os.system(f'open "{LOG_FILE_PATH}"')
    else:
        label.config(text=f"Unsupported system: {system_name}")

def log_to_excel(start_datetime, end_datetime, session_seconds):
    # Prepare values
    date_str = start_datetime.strftime("%d %B %Y")
    day_str = start_datetime.strftime("%A")
    start_str = start_datetime.strftime("%H:%M:%S")
    end_str = end_datetime.strftime("%H:%M:%S")

    # Deduct 1-hour break
    session_seconds -= 3600
    session_hours = seconds_to_decimal_hours(session_seconds)

    # Load or create workbook
    if not os.path.exists(EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Log"
        ws.append(["Day", "Date", "Start Time", "End Time", "Total Hours Worked"])
    else:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

    # Append new row
    ws.append([day_str, date_str, start_str, end_str, session_hours])
    wb.save(EXCEL_FILE_PATH)


# === Timer Logic ===
def update_timer():
    elapsed = time.time() - start_time
    remaining = WORKING_HOURS - elapsed

    progress_var.set((elapsed / WORKING_HOURS) * 100)

    if remaining <= 0:
        label.config(text="Time's up!", fg="red")
        root.update()
        stop_timer()
    elif remaining <= WARNING_TIME:
        label.config(
            text=f"Warning: {format_time(remaining)} left! Save your work.",
            fg="orange",
        )
        root.after(1000, update_timer)
    else:
        remaining_display = format_time(remaining)
        label.config(
            text=f"Started at: {start_time_str}\nTime Remaining: {remaining_display}",
            fg="green",
        )
        root.after(1000, update_timer)


def stop_timer():
    if messagebox.askokcancel("Quit", "End this session?"):
        end_time = datetime.now()
        end_time_str = end_time.strftime("%H:%M:%S")
        session_seconds = time.time() - start_time

        # Log to text file
        log_event(f"Session Ended: {end_time_str} on {end_time.strftime('%A, %B %d, %Y')}")
        log_event(f"Total Hours Worked (excluding 1-hour break): {format_time(session_seconds - 3600)}")
        log_event("-" * 40)

        # Log to Excel
        log_to_excel(datetime.fromtimestamp(start_time), end_time, session_seconds)

        open_log_file()
        root.quit()


# === GUI Setup ===
log_event(f"Session Started: {start_time_str} on {start_date_str}")

root = tk.Tk()
root.title("Work Timer")
root.geometry("400x200")
root.configure(bg="#282c34")

label = tk.Label(
    root, text="Time Remaining: ", font=("Helvetica", 16), fg="white", bg="#282c34"
)
label.pack(expand=True)

progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(
    root, orient="horizontal", length=300, mode="determinate", variable=progress_var
)
progress_bar.pack(pady=10)

exit_button = tk.Button(
    root,
    text="Exit",
    command=stop_timer,
    bg="#61afef",
    fg="white",
    font=("Helvetica", 12),
    relief="flat",
)
exit_button.pack(side="bottom", pady=10)

update_timer()
root.mainloop()
