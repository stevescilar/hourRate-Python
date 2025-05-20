import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
file1 = "Active Proposal_Pipelines 5-15-2025 3-45-03 PM.xlsx"
file2 = "ZA_Pipeline 15_05_2025.xlsx"
output_file = "ZA_Pipeline_Highlighted_missing.xlsx"

# Load data
df1 = pd.read_excel(file1, engine="openpyxl")
df2 = pd.read_excel(file2, engine="openpyxl")

# Extract OpportunityIDs
ids_file1 = df1["OpportunityID"].dropna().astype(str)
ids_file2 = df2["OpportunityID"].dropna().astype(str)

# Find missing IDs
missing_ids = ids_file1[~ids_file1.isin(ids_file2)]

# Save a copy of the original for highlighting
df1.to_excel(output_file, index=False)

# Load workbook to apply formatting
wb = load_workbook(output_file)
ws = wb.active

# Highlight style
highlight = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)  # Yellow

# Find the 'OpportunityID' column index
id_col_index = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == "OpportunityID":
        id_col_index = col[0].column
        break

# Highlight missing IDs
if id_col_index:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=id_col_index)
        if str(cell.value) in missing_ids.values:
            cell.fill = highlight

# Save highlighted file
wb.save(output_file)

print(f"Done! Highlighted missing IDs in: {output_file}")
